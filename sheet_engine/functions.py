import pandas as pd
from pathlib import Path
from django.db.models import Q
from openpyxl import load_workbook
from xlrd.biffh import XLRDError
from django.db import IntegrityError
from staff.models import Staff, HouseMaster
from logger.utils import log_system_error
from students.models import (Subject,
                             Course,
                             Klass,
                             Student,
                             TeacherClassSubjectCombination,
                             Record,
                             ClassTeacherRemark,
                             HouseMasterRemark)

STUDENT_SHEET_START_ROW = 11
SHEET_SKIP_ROWS = 10
BASE_DIR = Path(__file__).resolve().parent.parent


def insert_subjects(url):
    try:
        data = pd.read_excel(url, sheet_name="subjects",
                             skiprows=SHEET_SKIP_ROWS,  dtype = str)
        df = pd.DataFrame(data)
        for index, row in df.iterrows():
            is_elective = 'y' in row['IS_ELECTIVE(Y/N)'].lower()
            subject, created = Subject.objects.get_or_create(
                subject_id=row['SUBJECT_ID'],
                name=row['NAME'],
                is_elective=is_elective
            )
    except (XLRDError) as e:
        return str(e)
    except Exception as e:
        log_system_error("insert_subjects", str(e))
        return False
    return True


def insert_courses(url):
    try:
        data = pd.read_excel(url, sheet_name="courses",
                             skiprows=SHEET_SKIP_ROWS,  dtype = str)
        df = pd.DataFrame(data)
        for index, row in df.iterrows():
            subjects = Subject.objects.filter(
                Q(subject_id__in=row['ELECTIVES_IDS'].strip().split()) |
                Q(is_elective=False)
            )
            course, created = Course.objects.get_or_create(
                course_id=row['COURSE_ID'].strip())
            course.name = row['NAME'].strip()
            course.subjects.add(*subjects)
            course.save()
    except (XLRDError, Subject.DoesNotExist) as e:
        return str(e)
    except Exception as e:
        log_system_error("insert_courses", str(e))
        return False
    return True


def insert_staff(url):
    try:
        data = pd.read_excel(url, sheet_name="staff",
                             skiprows=SHEET_SKIP_ROWS,  dtype = str)
        df = pd.DataFrame(data)
        for index, row in df.iterrows():
            staff, created = Staff.objects.get_or_create(
                staff_id=row['STAFF_ID'])
            staff.surname = row['SURNAME']
            staff.other_names = row['OTHER_NAMES']
            staff.sms_number = row['SMS_NUMBER']
            staff.has_class = "y" in row['HAS_CLASS'].lower()
            staff.gender = row['GENDER']
            staff.save()
    except (XLRDError, IntegrityError) as e:
        return str(e)
    except Exception as e:
        log_system_error("insert_staff", str(e))
        return False
    return True


def insert_classes(url):
    try:
        data = pd.read_excel(url, sheet_name="classes",
                             skiprows=SHEET_SKIP_ROWS,  dtype = str)
        df = pd.DataFrame(data)
        for index, row in df.iterrows():
            course = Course.objects.get(course_id=row['COURSE_ID'])
            class_teacher = Staff.objects.get(
                staff_id=row['CLASS_TEACHER_ID'])
            klass, created = Klass.objects.get_or_create(
                class_id=row['CLASS_ID'],
                name=row['NAME'],
                form=row['FORM'],
                stream=row['STREAM'],
                course=course
            )
            klass.class_teacher = class_teacher
            klass.save()
    except (XLRDError, IntegrityError, Course.DoesNotExist, Staff.DoesNotExist) as e:
        error_message = f"Row: {index+1}: "+str(e)
        return error_message
    except Exception as e:
        log_system_error("insert_classes", f"Error: Row: {index+1}: "+str(e))
        return False
    return True


def insert_students(url):
    try:
        data = pd.read_excel(url, sheet_name="students",
                             skiprows=SHEET_SKIP_ROWS,  dtype = str)
        df = pd.DataFrame(data)
        for index, row in df.iterrows():
            student_id = row['STUDENT_ID']
            klass = Klass.objects.get(class_id=row['CLASS_ID'])
            electives = klass.course.subjects.filter(
                subject_id__in=row['ELECTIVE_SUBJECT_IDS'].strip().split(),
                is_elective=True
            )
            if electives.count() != len(row['ELECTIVE_SUBJECT_IDS'].strip().split()):
                error_message = f"Error: Student ID: {student_id}: Course  and Subject mismatch."
                return error_message

            student, created = Student.objects.get_or_create(
                student_id=student_id,
                surname=row['SURNAME'],
                other_names=row['OTHER_NAMES'],
                sms_number=row['SMS_NUMBER']
            )
            student.electives.add(*electives)
            student.house = row['HOUSE']
            student.track = row['TRACK']
            student.bio = row['BIO']
            student.gender = row['GENDER']
            student.father = row['FATHER']
            student.mother = row['MOTHER']
            student.klass = klass
            student.save()
    except (XLRDError, IntegrityError, Subject.DoesNotExist, Klass.DoesNotExist) as e:
        error_message = f"Error: Student ID: {student_id}: "+str(e)
        return error_message
    except Exception as e:
        log_system_error("insert_students",
                         f"Error: Student ID: {student_id}: "+str(e))
        return False
    return True


def insert_house_masters(url):
    try:
        data = pd.read_excel(url, sheet_name="house_masters",
                             skiprows=SHEET_SKIP_ROWS,  dtype = str)
        df = pd.DataFrame(data)
        for index, row in df.iterrows():
            staff = Staff.objects.get(staff_id=row["STAFF_ID"])
            HouseMaster.objects.create(house=row["HOUSE"], staff=staff)
    except (XLRDError, Staff.DoesNotExist) as e:
        error_message = f"Row: {index+1}: "+str(e)
        return error_message
    except Exception as e:
        log_system_error("insert_house_masters",
                         f"Error: Row: {index+1}: "+str(e))
        return False
    return True


def generate_subject_staff_combinations():
    url = BASE_DIR / "sheet_engine/static/template_sheets/teacher_subjects.xlsx"
    workbook = load_workbook(filename=url)
    worksheet = workbook["teacher_subjects"]
    subjects = Subject.objects.all()
    column_index = 1
    row_index = 12
    for subject in subjects:
        for course in subject.courses.all():
            for klass in course.classes.all():
                combination = TeacherClassSubjectCombination.objects.filter(
                    subject = subject,
                    klass = klass,
                ).first()
                worksheet.cell(column=column_index, row=row_index,
                               value=subject.subject_id)
                worksheet.cell(column=column_index+1,
                               row=row_index, value=subject.name)
                worksheet.cell(column=column_index+2,
                               row=row_index, value=klass.class_id)
                worksheet.cell(column=column_index+3,
                               row=row_index, value=klass.name)
                if combination:
                    worksheet.cell(column=column_index+4,
                                row=row_index, value=combination.staff.staff_id)
                    worksheet.cell(column=column_index+5,
                                row=row_index, value=combination.staff.get_full_name())
                row_index += 1
    workbook.save(BASE_DIR / "sheet_engine/static/generated_sheets/generate_subject_staff_combinations.xlsx")
    return "/static/generated_sheets/generate_subject_staff_combinations.xlsx"



def insert_subject_staff_combination(url):
    op_index = 0
    try:
        data = pd.read_excel(url, sheet_name="teacher_subjects",
                             skiprows=SHEET_SKIP_ROWS,  dtype = str)
        df = pd.DataFrame(data)
        for index, row in df.iterrows():
            op_index = index
            subject = Subject.objects.get(subject_id=row['SUBJECT_ID'])
            klass = Klass.objects.get(class_id=row['CLASS_ID'])
            staff = Staff.objects.get(staff_id=row['STAFF_ID'])
            TeacherClassSubjectCombination.objects.get_or_create(
                subject=subject, klass=klass, staff=staff)
    except (XLRDError, Subject.DoesNotExist, Klass.DoesNotExist, Staff.DoesNotExist) as e:
        error_message = f"Row: {op_index+1}: "+str(e)
        return error_message
    except Exception as e:
        log_system_error("insert_subject_staff_combination",
                         f"Error: Row: {op_index+1}: "+str(e))
        return False
    return True


def generate_record_sheet(staff_id, academic_year, semester, form, subject_id):
    url = r"D:\CodeLab\Projects\curie\sheet_engine\test_sheets\academic_records.xlsx"
    url = BASE_DIR/"sheet_engine/template_sheets/academic_records.xlsx"
    column_index = 1
    row_index = 12
    try:
        workbook = load_workbook(filename=url)
        worksheet = workbook["records"]

        staff = Staff.objects.get(staff_id=staff_id)
        subject = Subject.objects.get(subject_id=subject_id)
        teaches = staff.teaches.filter(subject=subject, klass__form=form)
        classes = [item.klass for item in teaches]
        students = Student.objects.filter(klass__in=classes)
        for index, student in enumerate(students):
            row_index += index
            worksheet.cell(column=column_index, row=row_index, value=staff_id)
            worksheet.cell(column=column_index+1,
                           row=row_index, value=academic_year)
            worksheet.cell(column=column_index+2,
                           row=row_index, value=semester)
            worksheet.cell(column=column_index+3, row=row_index,
                           value=student.klass.class_id)
            worksheet.cell(column=column_index+4, row=row_index,
                           value=subject.subject_id)
            worksheet.cell(column=column_index+5, row=row_index,
                           value=student.student_id)
            worksheet.cell(column=column_index+6,
                           row=row_index, value=f"{student.surname} {student.other_names}")
      
        workbook.save(BASE_DIR / "sheet_engine/static/generated_sheets/generate_record_sheet.xlsx")   
        # workbook.save(
        #     r"D:\CodeLab\Projects\curie\sheet_engine\generated_sheets\generate_record_sheet.xlsx")
    except (XLRDError, Subject.DoesNotExist, Staff.DoesNotExist) as e:
        error_message = f"Row: {index+1}: "+str(e)
        return error_message
    except Exception as e:
        log_system_error("generate_record_sheet",
                         f"Error: Row: {index+1}: "+str(e))
        return False
    return True


def insert_records(url):
    try:
        data = pd.read_excel(url, sheet_name="records",
                             skiprows=SHEET_SKIP_ROWS,  dtype = str)
        df = pd.DataFrame(data)
        for index, row in df.iterrows():
            subject = Subject.objects.get(subject_id=row["SUBJECT_ID"])
            student = Student.objects.get(student_id=row["STUDENT_ID"])
            klass = Klass.objects.get(class_id=row["CLASS_ID"])

            record, created = Record.objects.get_or_create(
                academic_year=row['ACADEMIC_YEAR'],
                semester=row['SEMESTER'],
                subject=subject,
                klass=klass,
                student=student,
            )
            record.class_score = row['CLASS_SCORE']
            record.exam_score = row['EXAM_SCORE']
            record.save()
    except (XLRDError, Subject.DoesNotExist, Student.DoesNotExist, Klass.DoesNotExist) as e:
        error_message = f"Row: {index+1}: "+str(e)
        return error_message
    except Exception as e:
        log_system_error("insert_records", f"Error: Row: {index+1}: "+str(e))
        return False
    return True


def generate_teacher_remark_sheet(class_id, academic_year, semester, total_attendance):
    url = BASE_DIR/"sheet_engine/template_sheets/class_teacher_remarks.xlsx"
    column_index = 1
    row_index = 12
    try:
        workbook = load_workbook(filename=url)
        worksheet = workbook["class_teacher_remarks"]
        klass = Klass.objects.get(class_id=class_id)
        students = Student.objects.filter(klass=klass)
        for index, student in enumerate(students):
            row_index += index
            worksheet.cell(column=column_index, row=row_index,
                           value=student.student_id)
            worksheet.cell(column=column_index+1, row=row_index,
                           value=student.get_full_name())
            worksheet.cell(column=column_index+2,
                           row=row_index, value=academic_year)
            worksheet.cell(column=column_index+3,
                           row=row_index, value=semester)
            worksheet.cell(column=column_index+4, row=row_index,
                           value=total_attendance)
        output_url = BASE_DIR / \
            f"sheet_engine/static/generated_sheets/{class_id}_{academic_year.replace('/','_')}_{semester}_remarks.xlsx"
        workbook.save(output_url)
    except (XLRDError, Klass.DoesNotExist) as e:
        error_message = f"Row: {index+1}: "+str(e)
        return error_message
    except Exception as e:
        log_system_error("generate_teacher_remark_sheet",
                         f"Error: Row: {index+1}: "+str(e))
        return False
    return True


def insert_teacher_remarks(url, class_id):
    try:
        data = pd.read_excel(url, sheet_name="class_teacher_remarks",
                             skiprows=SHEET_SKIP_ROWS,  dtype = str)
        df = pd.DataFrame(data)
        for index, row in df.iterrows():
            student = Student.objects.get(student_id=row['STUDENT_ID'])
            klass = Klass.objects.get(class_id=class_id)
            ClassTeacherRemark.objects.get_or_create(
                student=student,
                klass=klass,
                academic_year=row['ACADEMIC_YEAR'],
                semester=row['SEMESTER'],
                total_attendance=row['TOTAL_ATTENDANCE'],
                attendance=row['ATTENDANCE'],
                attitude=row['ATTITUDE'],
                interest=row['INTEREST'],
                conduct=row['CONDUCT'],
                remark=row['REMARK']
            )
    except (XLRDError, Student.DoesNotExist, Klass.DoesNotExist) as e:
        error_message = f"Row: {index+1}: "+str(e)
        return error_message
    except Exception as e:
        log_system_error("insert_teacher_remarks",
                         f"Error: Row: {index+1}: "+str(e))
        return False
    return True


def generate_house_master_remark_sheet(academic_year, semester, house):
    url = BASE_DIR/"sheet_engine/template_sheets/house_master_remarks.xlsx"
    column_index = 1
    row_index = 12
    try:
        workbook = load_workbook(filename=url)
        worksheet = workbook["house_master_remarks"]
        students = Student.objects.filter(house=house)
        for index, student in enumerate(students):
            row_index += index
            worksheet.cell(column=column_index, row=row_index,
                           value=academic_year)
            worksheet.cell(column=column_index+1, row=row_index,
                           value=semester)
            worksheet.cell(column=column_index+2,
                           row=row_index, value=student.student_id)
            worksheet.cell(column=column_index+3,
                           row=row_index, value=student.get_full_name())
        output_url = BASE_DIR / \
            f"sheet_engine/static/generated_sheets/{house}_{academic_year.replace('/','_')}_{semester}_remarks.xlsx"
        workbook.save(output_url)
    except (XLRDError) as e:
        error_message = f"Row: {index+1}: "+str(e)
        return error_message
    except Exception as e:
        log_system_error("generate_house_master_remark_sheet",
                         f"Error: Row: {index+1}: "+str(e))
        return False
    return True


def insert_house_master_remarks(url):
    try:
        data = pd.read_excel(url, sheet_name="house_master_remarks",
                             skiprows=SHEET_SKIP_ROWS,  dtype = str)
        df = pd.DataFrame(data)
        for index, row in df.iterrows():
            student = Student.objects.get(student_id=row['STUDENT_ID'])
            HouseMasterRemark.objects.get_or_create(
                student=student,
                klass=student.klass,
                academic_year=row['ACADEMIC_YEAR'],
                semester=row['SEMESTER'],
                remark=row['REMARK']
            )
    except (XLRDError, Student.DoesNotExist) as e:
        error_message = f"Row: {index+1}: "+str(e)
        return error_message
    except Exception as e:
        log_system_error("insert_house_master_remarks",
                         f"Error: Row: {index+1}: "+str(e))
        return False
    return True
